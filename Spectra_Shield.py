import os
import sys
import csv
import copy
import json
import time
import pytz
import shutil
import getpass
import argparse
import textwrap
import requests
import colorama
import openpyxl
import subprocess
import prettytable
from tzlocal import get_localzone
from colorama import Fore, Style, init
from prettytable import PrettyTable, ALL
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from alive_progress import alive_bar, config_handler

colorama.init(autoreset=True)

def print_banner():
    banner = f"""
{Fore.CYAN}╔════════════════════════════════════════════════════════════════════╗
║ {Fore.YELLOW}  ____                  _              ____  _     _      _     _ {Fore.CYAN} ║
║ {Fore.YELLOW} / ___| _ __   ___  ___| |_ _ __ __ _ / ___|| |__ (_) ___| | __| |{Fore.CYAN} ║
║ {Fore.YELLOW} \___ \| '_ \ / _ \/ __| __| '__/ _` |\___ \| '_ \| |/ _ \ |/ _` |{Fore.CYAN} ║
║ {Fore.YELLOW}  ___) | |_) |  __/ (__| |_| | | (_| | ___) | | | | |  __/ | (_| |{Fore.CYAN} ║
║ {Fore.YELLOW} |____/| .__/ \___|\___|\__|_|  \__,_||____/|_| |_|_|\___|_|\__,_|{Fore.CYAN} ║
║ {Fore.YELLOW}       |_|                                                        {Fore.CYAN} ║
╚════════════════════════════════════════════════════════════════════╝
{Fore.WHITE}v1.0

{Fore.YELLOW}\033[1mAuthor: Himanshu Kumar\033[0m
{Fore.LIGHTBLUE_EX}https://www.linkedin.com/in/himanshuk8/
{Fore.LIGHTBLUE_EX}https://www.github.com/himanshughub
    """
    print(banner)
    print(f"{Fore.WHITE}{Style.BRIGHT}Welcome to Spectra_Shield - Protecting Your Digital Spectrum{Style.RESET_ALL}")
    print(f"{Fore.YELLOW}{Style.BRIGHT}About Spectra_Shield:{Style.RESET_ALL}")
    print(f"{Fore.WHITE}Spectra_Shield is an advanced open-source threat hunting tool designed to enhance{Style.RESET_ALL}")
    print(f"{Fore.WHITE}threat detection capabilities in cybersecurity operations. It is specifically{Style.RESET_ALL}")
    print(f"{Fore.WHITE}engineered to cover loopholes of CrowdStrike where its threat intelligence{Style.RESET_ALL}")
    print(f"{Fore.WHITE}database is not comprehensive to find various known malicious processes.{Style.RESET_ALL}")
    print(f"{Fore.WHITE}By leveraging multiple threat intelligence sources, including VirusTotal API,{Style.RESET_ALL}")
    print(f"{Fore.WHITE}Abuse.ch API, and custom threat intelligence databases, Spectra_Shield provides{Style.RESET_ALL}")
    print(f"{Fore.WHITE}a comprehensive solution for identifying malicious processes executed on machines,{Style.RESET_ALL}")
    print(f"{Fore.WHITE}filling critical gaps in existing security infrastructures.{Style.RESET_ALL}")

    print(f"\n{Fore.YELLOW}{Style.BRIGHT}Key Features:{Style.RESET_ALL}")
    print(f"•{Fore.WHITE} Integration with VirusTotal API for extensive threat analysis{Style.RESET_ALL}")
    print(f"•{Fore.WHITE} Utilization of Abuse.ch API for up-to-date threat intelligence{Style.RESET_ALL}")
    print(f"•{Fore.WHITE} Custom hash list investigation for tailored threat detection{Style.RESET_ALL}")
    print(f"•{Fore.WHITE} Full scan capability combining all available investigation methods{Style.RESET_ALL}")

def VirusTotal_API_Investigate(skip_prompt=False):
    if not skip_prompt:
        print(f"{Fore.RED}{Style.BRIGHT}WARNING{Style.RESET_ALL}{Fore.WHITE}: This option uses the VirusTotal open source API, which may have terms of service restrictions for business use.")
        while True:
            confirm = input(f"\033[1;33mAre you sure you want to proceed with VirusTotal option? (yes/no): \033[0m").lower()
            if confirm == 'yes':
                break
            elif confirm == 'no':
                print("Returning to main menu.")
                return
            else:
                print("Invalid input. Please enter 'yes' or 'no'.")

    print("Proceeding with VirusTotal API investigation...")
    # Add your code for VirusTotal API investigation here

    def get_masked_api_key():
        """Get a single API key with masked input"""
        try:
            key = getpass.getpass(f"{Fore.YELLOW}Enter VirusTotal API key (input will be hidden): {Style.RESET_ALL}")
            return key.strip()
        except (KeyboardInterrupt, EOFError):
            print("\nInput cancelled by user.")
            return None

    def get_api_keys():
        """Get API keys from command line arguments or prompt user for input"""
        parser = argparse.ArgumentParser(description='VirusTotal API Key Handler')
        parser.add_argument('--api-keys', nargs='*', help='One or more VirusTotal API keys')
        args = parser.parse_args()

        api_keys = []

        # If API keys were provided as command line arguments
        if args.api_keys:
            api_keys.extend(args.api_keys)

        # If no API keys were provided through arguments, prompt user
        if not api_keys:
            print(f"{Fore.YELLOW}Please enter your VirusTotal API key(s).{Style.RESET_ALL}")
            print(f"{Fore.YELLOW}Press Enter with empty input when done.{Style.RESET_ALL}")

            while True:
                key = get_masked_api_key()
                if not key:  # Empty line or cancelled input
                    if api_keys:  # If we already have some keys, break
                        break
                    else:
                        continue  # If no keys yet, continue prompting

                api_keys.append(key)
                print(f"{Fore.GREEN}API key added successfully!{Style.RESET_ALL}")
                print(f"{Fore.YELLOW}Enter another API key or press Enter to continue...{Style.RESET_ALL}")

        if not api_keys:
            print(f"{Fore.RED}No API keys provided. Exiting...{Style.RESET_ALL}")
            sys.exit(1)

        print(f"{Fore.GREEN}Successfully loaded {len(api_keys)} API key(s){Style.RESET_ALL}")
        return api_keys

    # VirusTotal v3 API Base URL
    base_url = "https://www.virustotal.com/api/v3/files/"

    # Get API keys at startup
    api_keys = get_api_keys()
    global current_api_key_index, dashboard_data
    current_api_key_index = 0

    def get_terminal_width():
        return shutil.get_terminal_size().columns

    def configure_progress_bar():
        config_handler.set_global(spinner="waves", bar="smooth", unknown="waves", length=72, enrich_print=False, force_tty=True, stats=False)

    def format_progress_text(malicious_hashes, eta):
        terminal_width = get_terminal_width()
        available_width = terminal_width - 80
        malicious_text = f"{Fore.RED}{Style.BRIGHT}{malicious_hashes}{Style.RESET_ALL}"
        eta_text = f"{int(eta)}s"
        progress_text = f"Malicious: {malicious_text} | ETA: {eta_text}"
        padded_text = progress_text.ljust(available_width)
        return padded_text[:available_width]

    def format_threat_classification(threat_classification):
        if not threat_classification:
            return "N/A"
        categories = ', '.join([f"{cat['value']} ({cat['count']})" for cat in threat_classification.get('popular_threat_category', [])])
        names = ', '.join([f"{name['value']} ({name['count']})" for name in threat_classification.get('popular_threat_name', [])])
        return f"Label: {threat_classification.get('suggested_threat_label', 'N/A')}\nCategories: {categories}\nNames: {names}"

    def format_timestamp(ts):
        if ts:
            return datetime.utcfromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')
        return "N/A"

    def get_current_timestamp():
        local_tz = get_localzone()
        current_time = datetime.now(local_tz)
        return current_time.strftime('%Y-%m-%d %H:%M:%S %Z')

    def check_hash(file_hash, file_name, computer_name, local_addresses, hash_count, total_hashes):
        global current_api_key_index, dashboard_data
        headers = {
            "accept": "application/json",
            "X-Apikey": api_keys[current_api_key_index]
        }
        url = base_url + file_hash
        vendor_count = "N/A"
        sign_info = "N/A"

        # Calculate the width needed for the counter
        counter_width = len(f"[{total_hashes}/{total_hashes}]")
        # Calculate the padding for all fields (counter width + 2 extra spaces)
        field_padding = counter_width + 2
        # Create the base padding
        base_padding = ' ' * field_padding

        # Function to print padded multiline content
        def print_padded(label, content, is_first_line=False):
            import textwrap
            label_width = len(label) + 2  # +2 for the colon and space
            total_padding = field_padding if is_first_line else field_padding + label_width

            # Fields that should not be bright
            non_bright_fields = ["File Name", "Last Analysis Stats", "SHA256", "MD5", "Status", "Details"]

            if label == "Popular Threat Classification":
                print(f"{' ' * field_padding}{Fore.YELLOW}{label}: {Style.RESET_ALL}", end="")
                if isinstance(content, dict):
                    padding = " " * (field_padding + len(label) + 2)  # +2 for ": "
                    first_item = True
                    for key, value in content.items():
                        if first_item:
                            print(f"{Fore.WHITE}{key}: {value}{Style.RESET_ALL}")
                            first_item = False
                        else:
                            print(f"{padding}{Fore.WHITE}{key}: {value}{Style.RESET_ALL}")
                elif isinstance(content, str):
                    lines = content.split('\n')
                    print(f"{Fore.WHITE}{lines[0]}{Style.RESET_ALL}")
                    padding = " " * (field_padding + len(label) + 2)  # +2 for ": "
                    for line in lines[1:]:
                        print(f"{padding}{Fore.WHITE}{line}{Style.RESET_ALL}")
                else:
                    print(f"{Fore.WHITE}{content}{Style.RESET_ALL}")

            else:
                # Apply BRIGHT style only if it's not in the non_bright_fields list
                style_bright = Style.BRIGHT if label not in non_bright_fields else ""
                wrapper = textwrap.TextWrapper(width=165, subsequent_indent=' ' * total_padding)
                wrapped_content = wrapper.wrap(str(content))
                print(f"{' ' * field_padding}{Fore.YELLOW}{style_bright}{label}: {Style.RESET_ALL}{Fore.WHITE}{wrapped_content[0]}{Style.RESET_ALL}")
                for line in wrapped_content[1:]:
                    print(f"{' ' * total_padding}{Fore.WHITE}{line}{Style.RESET_ALL}")

        try:
            response = requests.get(url, headers=headers)
            if response.status_code == 200:
                data = response.json()
                file_info = data.get('data', {})
                attributes = file_info.get('attributes', {})

                # Format the counter
                counter = f"[{hash_count}/{total_hashes}]"

                # Print the "Results for hash:" line with proper alignment
                print(f"\n{counter:<{field_padding}}{Fore.YELLOW}{Style.BRIGHT}Results for hash: {Style.RESET_ALL}{Fore.BLUE}{file_hash}{Style.RESET_ALL}")
                threat_classification = format_threat_classification(attributes.get('popular_threat_classification', {}))
                print_padded("Popular Threat Classification", threat_classification, is_first_line=True)
                print_padded("File Name", file_name)
                signature_info = attributes.get('signature_info', {})
                verified = signature_info.get('verified', 'N/A')
                product = signature_info.get('product', 'N/A')
                cert_issuer = next((cert.get('cert issuer', 'N/A') for cert in signature_info.get('x509', []) if 'cert issuer' in cert), 'N/A')
                if verified == 'N/A' and product == 'N/A' and cert_issuer == 'N/A':
                    sign_info = 'N/A'
                else:
                    sign_info = f"{verified}, {product}, {cert_issuer}"

                last_analysis_stats = attributes.get('last_analysis_stats', {})
                stats_items = []
                for key, value in last_analysis_stats.items():
                    if key == 'malicious' and value > 0:
                        stats_items.append(f"{Fore.RED}{Style.BRIGHT}'{key}': {value}{Style.RESET_ALL}")
                    else:
                        stats_items.append(f"{Fore.WHITE}'{key}': {value}{Style.RESET_ALL}")

                stats_string = "{ " + ", ".join(stats_items) + " }"
                print_padded("Last Analysis Stats", stats_string)

                print_padded("SHA256", attributes.get('sha256', 'N/A'))
                print_padded("MD5", attributes.get('md5', 'N/A'))

                malicious = last_analysis_stats.get('malicious', 0)
                suspicious = last_analysis_stats.get('suspicious', 0)
                vendor_count = f"{malicious + suspicious}/{sum(last_analysis_stats.values()) - last_analysis_stats.get('type-unsupported', 0)}"
                if malicious > 0 or suspicious > 0:
                    indicator = "malicious" if malicious > 0 else "suspicious"
                    existing_entry = next((item for item in dashboard_data if item['hash'] == file_hash), None)
                    if existing_entry:
                        existing_entry['file_names'].add(file_name)
                        existing_entry['computer_names'].add(computer_name)
                        existing_entry['local_addresses'].update(local_addresses)
                    else:
                        dashboard_data.append({
                            'hash': file_hash,
                            'md5': attributes.get('md5', 'N/A'),
                            'indicator': indicator,
                            'file_names': {file_name},
                            'computer_names': {computer_name},
                            'local_addresses': set(local_addresses),
                            'vendor_count': vendor_count,
                            'threat_classification': attributes.get('popular_threat_classification', {}),
                            'sign_info': sign_info
                        })
            elif response.status_code == 404:
                # Format the counter
                counter = f"[{hash_count}/{total_hashes}]"

                # Print the "Results for hash:" line with proper alignment
                print(f"\n{counter:<{field_padding}}{Fore.YELLOW}{Style.BRIGHT}Results for hash: {Style.RESET_ALL}{Fore.BLUE}{file_hash}{Style.RESET_ALL}")
                print_padded("Status", "No match found", is_first_line=True)
                print_padded("Details", "The provided hash was not found in the VirusTotal database.")
                # Add to dashboard with 'Unknown' indicator
                dashboard_data.append({
                    'hash': file_hash,
                    'indicator': 'unknown',
                    'file_names': {file_name},
                    'computer_names': {computer_name},
                    'local_addresses': set(local_addresses),
                    'vendor_count': vendor_count,
                    'sign_info': 'N/A'
                })
            else:

                # Format the counter
                counter = f"[{hash_count}/{total_hashes}]"
                # Print the "Results for hash:" line with proper alignment
                print(f"\n{counter:<{field_padding}}{Fore.YELLOW}{Style.BRIGHT}Results for hash: {Style.RESET_ALL}{Fore.BLUE}{file_hash}{Style.RESET_ALL}")
                print_padded("Status", f"Error (HTTP {response.status_code})", is_first_line=True)
                print_padded("Details", response.text)
        except requests.exceptions.ConnectionError:
            sad_face = f"""
            {Fore.RED}{Style.BRIGHT}
            (╯︵╰)
            {Style.RESET_ALL}
            """
            print(f"{sad_face}{Fore.RED}{Style.BRIGHT}Error{Style.RESET_ALL}: No internet connection detected. Please check your network connection and try again.\n")
            exit()

        # Move to the next API key in the pool
        current_api_key_index = (current_api_key_index + 1) % len(api_keys)

    def generate_filename():
        edt = pytz.timezone('US/Eastern')
        now = datetime.now(edt)
        return f"SpectraShield_Dashboard.xlsx"

    def find_first_available_row(worksheet):
        for row in range(1, worksheet.max_row + 2):
            if all(cell.value is None for cell in worksheet[row]):
                return row
        return worksheet.max_row + 1

    def sort_excel_sheet(ws):
        # custom sorting order for 'Indicator' column
        indicator_order = {'Malicious': 1, 'Suspicious': 2, 'Unknown': 3}
        # custom sorting order for 'Severity' column
        severity_order = {'Critical': 1, 'High': 2, 'Medium': 3, 'Low': 4}
        # custom sorting order for 'Status' column
        status_order = {'OPEN': 1, 'WORK_IN_PROGRESS': 2, 'SOC_VALIDATION': 3, 'WAITING_FOR_CUSTOMER': 4, 'CLOSED': 5}
        # Get all rows except the header
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))

        def safe_timestamp(date_string):
            if date_string and isinstance(date_string, str):
                try:
                    return -datetime.strptime(date_string, '%Y-%m-%d %H:%M:%S %Z').timestamp()
                except ValueError:
                    # If the date string doesn't match the expected format, return a default value
                    return float('inf')  # sort to the end
            return float('inf')  # sort None or empty values to the end

        def safe_get(dictionary, key, default):
            if dictionary is None:
                return default
            return dictionary.get(key, default)

        # Sort the rows
        sorted_rows = sorted(
            data_rows,
            key=lambda x: (
                status_order.get(x[0], 6) if x[0] else 6,  # Sort by Status (column 1), empty cells to the end
                safe_timestamp(x[3]),  # Sort by Report_timestamp (column 4) in descending order
                severity_order.get(x[7], 5) if x[7] else 6,  # Sort by Severity (column 8)
                indicator_order.get(x[5], 4) if x[5] else 5,  # Sort by Indicator (column 6)
            )
        )

        # Clear the existing data (except header)
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None

        # Write the sorted data back to the sheet
        for row_index, row_data in enumerate(sorted_rows, start=2):
            for col_index, cell_value in enumerate(row_data, start=1):
                ws.cell(row=row_index, column=col_index, value=cell_value)

        # Re-apply styles after sorting
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.column in [1, 2, 3, 4, 5, 6, 7, 8, 9]:
                    cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                # Re-apply hyperlink and font for the "Source" column
                if cell.column == 9:
                    cell.font = Font(color="ADD8E6", underline="single")
                    source_value = cell.value
                    sha256_hash = ws.cell(row=cell.row, column=12).value
                    if source_value == "abuse.ch":
                        cell.hyperlink = "https://bazaar.abuse.ch/browse/"
                    elif sha256_hash:
                        cell.hyperlink = f"https://www.virustotal.com/gui/file/{sha256_hash}"

        # Re-apply data validation for the "Status" column
        status_options = ["OPEN", "WORK_IN_PROGRESS", "SOC_VALIDATION", "WAITING_FOR_CUSTOMER", "CLOSED"]
        status_dv = DataValidation(type="list", formula1=f'"{",".join(status_options)}"', allow_blank=False)
        ws.add_data_validation(status_dv)
        for cell in ws['A'][2:]:  # Status is column A, start from row 2
            status_dv.add(cell)

        # Re-apply data validation for the "Verdict" column
        verdict_options = ["True Positive", "False Positive", "Indeterminate"]
        dv = DataValidation(type="list", formula1=f'"{",".join(verdict_options)}"', allow_blank=True)
        ws.add_data_validation(dv)
        for cell in ws['C'][2:]:  #Verdict is column C
            dv.add(cell)

    def find_existing_verdict(ws, sha256_hash):
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[11] == sha256_hash and row[2]:  # SHA256HashData is in column 12 (index 11), Verdict is in column 3 (index 2)
                return row[2]
        return None

    start_time = time.time()

    # Initialize dashboard data
    dashboard_data = []

    # Get the CSV file path
    while True:
        file_path = input(f"\033[1;33mPlease enter the full path to your CSV file: \033[0m").strip()
        if os.path.isfile(file_path):
            break
        else:
            print("The file does not exist. Please check the path and try again.")

    # Count total number of hashes
    print("Counting total hashes...")
    with open(file_path, 'r') as csvfile:
        csvreader = csv.DictReader(csvfile)
        total_hashes = sum(1 for row in csvreader if row.get('SHA256HashData'))

    print(f"Total hashes to process: {total_hashes}")

    # Process hashes
    print("Processing hashes...")
    malicious_hashes = 0
    start_time = time.time()

    configure_progress_bar()

    with alive_bar(total_hashes, title="Processing hashes", enrich_print=False) as bar:
        with open(file_path, 'r') as csvfile:
            csvreader = csv.DictReader(csvfile)
            for hash_count, row in enumerate(csvreader, 1):
                file_hash = row.get('SHA256HashData')
                if not file_hash:
                    continue
                file_name = row.get('FileName', 'N/A')
                computer_name = row.get('ComputerName', 'N/A')
                local_addresses = row.get('LocalAddressIP4', '').split(';')
                local_addresses = [ip.strip() for ip in local_addresses if ip.strip()]
                aid = row.get('aid', 'N/A')
                cid = row.get('cid', 'N/A')
                file_path_value = row.get('FilePath', 'N/A')

                check_hash(file_hash, file_name, computer_name, local_addresses, hash_count, total_hashes)

                # Update dashboard_data and count malicious hashes
                for item in dashboard_data:
                    if item['hash'] == file_hash:
                        item['aid'] = aid
                        item['cid'] = cid
                        item['file_path'] = file_path_value
                        if item['indicator'] in ['malicious', 'suspicious']:
                            malicious_hashes += 1
                        break

                # Calculate elapsed time and estimate the remaining time
                elapsed_time = time.time() - start_time
                estimated_total_time = (elapsed_time / hash_count) * total_hashes
                eta = estimated_total_time - elapsed_time

                # Update the progress bar with malicious hashes and ETA
                progress_text = format_progress_text(malicious_hashes, eta)
                bar.text(progress_text)
                bar()

                time.sleep(0)  # Wait for a short time between requests to avoid hitting API rate limits

    # After processing all hashes, export to Excel and then display the dashboard
    if dashboard_data:

        # Get the current timestamp
        report_timestamp = get_current_timestamp()

        # Sort dashboard_data based on indicator priority
        indicator_priority = {"malicious": 0, "suspicious": 1, "unknown": 2}
        dashboard_data.sort(key=lambda x: indicator_priority[x['indicator']])

        # Check if the file already exists
        filename = generate_filename()

        if os.path.exists(filename):
            # Load existing workbook
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
        else:
            # Create new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Dashboard"

        # Write headers
        headers = ["Status", "Assigned_to", "Verdict", "Report_timestamp", "sign_info", "Indicator", "Vendors Reported", "Severity", "Source", "ThreatClassification", "MD5HashData", "SHA256HashData", "FileName", "ComputerName", "LocalAddressIP4", "aid", "cid", "FilePath", "Additional Notes"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        start_row = find_first_available_row(ws)

        # Create data validation for Verdict column
        verdict_options = ["True Positive", "False Positive", "Indeterminate"]
        dv = DataValidation(type="list", formula1=f'"{",".join(verdict_options)}"', allow_blank=True)
        ws.add_data_validation(dv)

        # Write data
        for row, item in enumerate(dashboard_data, start=start_row):
            malicious_str = item['vendor_count'].split('/')[0]
            malicious = int(malicious_str) if malicious_str.isdigit() else 0 
            suspicious = sum(1 for _ in item['vendor_count'] if 'suspicious' in _)

            # Determine severity
            if item.get('sign_info', 'N/A') != 'N/A':
                severity = "Medium"
            elif malicious >= 5:
                severity = "Critical"
            elif malicious > 0:
                severity = "High"
            elif suspicious > 0:
                severity = "Medium"
            else:
                severity = "Low"

            # Generate the VirusTotal link
            sha256_hash = item['hash']
            source_link = f"https://www.virustotal.com/gui/file/{sha256_hash}"

            ws.cell(row=row, column=1, value="OPEN")
            ws.cell(row=row, column=2, value="")  # Assigned_to

            existing_verdict = find_existing_verdict(ws, sha256_hash)
            if existing_verdict:
                ws.cell(row=row, column=3, value=existing_verdict)
            else:
                verdict_cell = ws.cell(row=row, column=3, value="")
                dv.add(verdict_cell)

            ws.cell(row=row, column=4, value=report_timestamp)
            ws.cell(row=row, column=5, value=item.get('sign_info', 'N/A'))
            ws.cell(row=row, column=6, value=item['indicator'].capitalize())
            ws.cell(row=row, column=7, value=item['vendor_count'])
            ws.cell(row=row, column=8, value=severity)

            # Insert "VirusTotal" text with hyperlink in the "Source" column
            source_value = item.get('source', 'VirusTotal')  # Default to VirusTotal if not specified
            source_cell = ws.cell(row=row, column=9, value=source_value)
            if source_value == "abuse.ch":
                source_cell.hyperlink = "https://bazaar.abuse.ch/browse/"
            else:
                source_cell.hyperlink = f"https://www.virustotal.com/gui/file/{sha256_hash}"
            source_cell.font = Font(color="ADD8E6", underline="single") 
            source_cell.alignment = Alignment(horizontal='center', vertical='top')

            threat_classification = format_threat_classification(item.get('threat_classification', {}))
            ws.cell(row=row, column=10, value=threat_classification)
            ws.cell(row=row, column=11, value=item.get('md5', 'N/A'))
            ws.cell(row=row, column=12, value=sha256_hash)
            ws.cell(row=row, column=13, value="\n".join(item['file_names']))
            ws.cell(row=row, column=14, value="\n".join(item['computer_names']))
            ws.cell(row=row, column=15, value="\n".join(item['local_addresses']))
            ws.cell(row=row, column=16, value=item.get('aid', ''))
            ws.cell(row=row, column=17, value=item.get('cid', ''))
            ws.cell(row=row, column=18, value=item.get('file_path', ''))
            ws.cell(row=row, column=19, value="") #empty column for Additional Notes

            # Set alignment for data cells
            for col in range(1, 20):
                cell = ws.cell(row=row, column=col)

                if col in [1, 2, 3, 4, 5, 6, 7, 8, 9, 19]:
                    cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # Dynamically adjust column widths and row heights
        for col in range(1, 18):
            column_letter = get_column_letter(col)
            max_length = 0
            column = ws[column_letter]
            for cell in column:
                try:
                    if cell.value:
                        # Calculate the maximum length for multi-line cell values
                        lines = str(cell.value).split('\n')
                        length = max(len(str(line).strip()) for line in lines)
                        if length > max_length:
                            max_length = length
                except:
                    pass

        def adjust_column_widths_and_row_heights(ws):
            # Define the light gray border
            light_gray_border = Border(
                left=Side(style='thin', color='444444'),
                right=Side(style='thin', color='444444'),
                top=Side(style='thin', color='444444'),
                bottom=Side(style='thin', color='444444')
            )

            # Define light gray 5 color
            light_gray_5 = "D3D3D3"

            # Find the column index for "FileName", "Assigned_to", and "Vendors Reported"
            filename_col = None
            assigned_to_col = None
            vendors_reported_col = None
            indicator_col = None
            source_col = None
            sign_info_col = None

            for col in range(1, ws.max_column + 1):
                header_value = ws.cell(row=1, column=col).value
                if header_value == "FileName":
                    filename_col = col
                elif header_value == "Assigned_to":
                    assigned_to_col = col
                elif header_value == "Vendors Reported":
                    vendors_reported_col = col
                elif header_value == "Indicator":
                    indicator_col = col
                elif header_value == "Source":
                    source_col = col
                elif header_value == "sign_info":
                    sign_info_col = col
                if filename_col and assigned_to_col and vendors_reported_col and indicator_col and source_col and sign_info_col:
                    break

            for col in range(1, ws.max_column + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                column = ws[column_letter]
                for cell in column:
                    try:
                        if cell.value:
                            new_alignment = Alignment(wrap_text=True, 
                                                horizontal=cell.alignment.horizontal, 
                                                vertical=cell.alignment.vertical)
                            cell.alignment = new_alignment
                            # Calculate the maximum length for multi-line cell values
                            lines = str(cell.value).split('\n')
                            length = max(len(str(line).strip()) for line in lines)
                            if length > max_length:
                                max_length = length

                        # Apply formatting only to non-header rows
                        if cell.row > 1:
                            # Set default font color to light gray 5
                            cell.font = Font(name='Calibri', size=11, color=light_gray_5)
                            # Special formatting for specific columns
                            if col == assigned_to_col:
                                cell.font = Font(name='Calibri', size=11, color="FFFFFF", bold=True)
                            elif col == filename_col:
                                cell.font = Font(name='Calibri', size=11, color="FFFF00")
                            elif col == vendors_reported_col and cell.value != "N/A":
                                cell.font = Font(name='Calibri', size=11, color="FF0000", bold=True)
                            elif col == indicator_col:
                                if cell.value == "Malicious":
                                    cell.font = Font(name='Calibri', size=11, color="FF0000", bold=True)  # Bold Red
                                elif cell.value == "Suspicious":
                                    cell.font = Font(name='Calibri', size=11, color="FFFF00", bold=True)  # Bold Yellow
                                elif cell.value == "Unknown":
                                    cell.font = Font(name='Calibri', size=11, color="ADD8E6", bold=True)  # Bold Light Blue
                            elif col == source_col:
                                cell.font = Font(name='Calibri', size=11, color=light_gray_5, underline="single")  # Underlined
                            elif col == sign_info_col and cell.value != "N/A":
                                cell.font = Font(name='Calibri', size=11, color="99FF99")  # Light Green 3

                        else:
                            # Set header font
                            cell.font = Font(name='Calibri', size=11, bold=True)
                        # Apply light gray border to all cells
                        cell.border = light_gray_border
                    except:
                        pass

                # Adjusting column width size
                if col == 1:  # Status column
                    adjusted_width = min(max(25, max_length + 4), 35)
                elif col == 2:  # Assigned_to
                    adjusted_width = max(25, max_length + 4)
                elif col == 3:  # Verdict
                    adjusted_width = max(20, max_length + 4)
                elif col == 4:  # Report_Timestamp
                    adjusted_width = max(18, max_length + 4)
                elif col == 5:  # sign_info column
                    adjusted_width = min(max(20, max_length + 4), 25)
                elif col == 9:  # Source
                    adjusted_width = min(max(15, max_length + 4), 20)
                elif col == 10:  # ThreatClassification
                    adjusted_width = min(max(40, max_length + 4), 50)
                elif col == 11:  # MD5HashData
                    adjusted_width = max(30, max_length + 4)
                elif col == 12:  # SHA256HashData
                    adjusted_width = max(65, max_length + 4)
                elif col == 13:  # FileName
                    adjusted_width = min(max(40, max_length + 4), 50)  # Cap at 50
                elif col == 14:  # ComputerName
                    adjusted_width = min(max(20, max_length + 4), 30)  # Cap at 30
                elif col == 15:  # LocalAddressIP4
                    adjusted_width = min(max(20, max_length + 4), 30)  # Cap at 30
                elif col == 16:  # aid column
                    adjusted_width = min(max(40, max_length + 4), 45)  # Cap at 40
                elif col == 17:  # cid column
                    adjusted_width = min(max(40, max_length + 4), 45)  # Cap at 40
                elif col == 18:  # FilePath column
                    adjusted_width = min(max(50, max_length + 4), 60)  # Cap at 60
                elif col == 19:  # Additional Notes
                    adjusted_width = 80
                else:
                    adjusted_width = min(max_length + 4, 45)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Add auto filter to all columns
            ws.auto_filter.ref = ws.dimensions

            # Adjust row heights, set wrap text for all cells, and apply black background
            black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            # Adjust row heights and set wrap text for all cells
            for row in ws.iter_rows():
                max_height = 14  # Default row height
                for cell in row:
                    if cell.value:
                        cell.alignment = Alignment(wrap_text=True,
                                                horizontal=cell.alignment.horizontal,
                                                vertical=cell.alignment.vertical)
                        lines = str(cell.value).split('\n')
                        row_height = 14 * len(lines)
                        if row_height > max_height:
                            max_height = row_height
                    # Apply black background to all cells except the header row
                    if cell.row > 1:
                        cell.fill = black_fill
                    cell.border = light_gray_border
                ws.row_dimensions[row[0].row].height = max_height

        # Sort the Excel sheet
        sort_excel_sheet(ws)

        adjust_column_widths_and_row_heights(ws)

        # Save the workbook
        filename = generate_filename()
        full_path = os.path.abspath(filename)
        wb.save(filename)

        total_hashes = len(dashboard_data)
        malicious_hashes = sum(1 for item in dashboard_data if item['indicator'] == 'malicious')

        end_time = time.time()
        elapsed_time = end_time - start_time

        # Display the dashboard on terminal
        print("\n" + "=" * 45)
        print(f"{Fore.CYAN}{Style.BRIGHT}Result Summary:{Style.RESET_ALL}")
        print(f"*{Fore.YELLOW}{Style.BRIGHT} Status:{Style.RESET_ALL} VirusTotal scan completed successfully")
        if malicious_hashes > 0:
            print(f"*{Fore.YELLOW}{Style.BRIGHT} Detection:{Style.RESET_ALL} {Fore.RED}{Style.BRIGHT}{malicious_hashes} Malicious Hash{'es' if malicious_hashes > 1 else ''} Detected!{Style.RESET_ALL}")
        else:
            print(f"*{Fore.YELLOW}{Style.BRIGHT} Detection:{Style.RESET_ALL} {Fore.GREEN}{Style.BRIGHT}No Malicious Hash Found!{Style.RESET_ALL}")
        print(f"*{Fore.YELLOW}{Style.BRIGHT} Scan option selected:{Style.RESET_ALL} VirusTotal API")
        print(f"*{Fore.YELLOW}{Style.BRIGHT} Result:{Style.RESET_ALL} {Fore.RED}{Style.BRIGHT}{malicious_hashes}{Style.RESET_ALL}/{total_hashes}")
        print(f"*{Fore.YELLOW}{Style.BRIGHT} Elapsed time:{Style.RESET_ALL} {elapsed_time:.2f} seconds")
        print("=" * 45)

        # Sort dashboard_data based on indicator priority
        indicator_priority = {"malicious": 0, "suspicious": 1, "unknown": 2}
        dashboard_data.sort(key=lambda x: indicator_priority[x['indicator']])

        def format_row(row):
            indicator, hash_data, file_names, computer_names, local_addresses = row

            # Define maximum width for each column
            max_width = {
                "FileName": 30,
                "ComputerName": 30,
                "LocalAddressIP4": 20
            }

            wrapped_file_names = "\n".join(["\n".join(textwrap.wrap(name, width=max_width["FileName"])) for name in file_names.split('\n')])
            wrapped_computer_names = "\n".join(["\n".join(textwrap.wrap(name, width=max_width["ComputerName"])) for name in computer_names.split('\n')])

            formatted_file_names = "\n".join([f"{Fore.YELLOW}{Style.BRIGHT}{name}{Style.RESET_ALL}" for name in wrapped_file_names.split('\n')])
            formatted_local_addresses = "\n".join([f"{Fore.GREEN}{Style.BRIGHT}{addr}{Style.RESET_ALL}" for addr in local_addresses.split('\n')])

            return [
                indicator,
                hash_data,
                formatted_file_names,
                f"{Style.RESET_ALL}{computer_names}",  # Explicitly reset style
                formatted_local_addresses
            ]

        table = PrettyTable()
        table.field_names = [
            f"{Fore.CYAN}{Style.BRIGHT}Indicator{Style.RESET_ALL}",
            f"{Fore.CYAN}{Style.BRIGHT}SHA256HashData{Style.RESET_ALL}",
            f"{Fore.CYAN}{Style.BRIGHT}FileName{Style.RESET_ALL}",
            f"{Fore.CYAN}{Style.BRIGHT}ComputerName{Style.RESET_ALL}",
            f"{Fore.CYAN}{Style.BRIGHT}LocalAddressIP4{Style.RESET_ALL}"
        ]

        # Set all column alignments to left
        for field in table.field_names:
            table.align[field] = "l"

        # Set header alignment to center
        table.header_align = "c"
        table.max_width["SHA256HashData"] = 65
        table.max_width["FileName"] = 40
        table.max_width["ComputerName"] = 30
        table.max_width["LocalAddressIP4"] = 20

        # Enable horizontal lines between rows
        table.hrules = prettytable.ALL

        for item in dashboard_data:
            indicator = item['indicator']
            if indicator == "malicious":
                indicator_color = Fore.LIGHTRED_EX
            elif indicator == "suspicious":
                indicator_color = Fore.LIGHTYELLOW_EX
            else:  # For 'unknown'
                indicator_color = Fore.LIGHTBLUE_EX

            indicator_text = f"{indicator_color}{Style.BRIGHT}●{Style.RESET_ALL} {indicator_color}{indicator.capitalize()}{Style.RESET_ALL}"

            # highlight filenames in bright yellow

            file_names = "\n".join(item['file_names'])
            computer_names = "\n".join(item['computer_names'])
            local_addresses = "\n".join(item['local_addresses'])

            row = [
                indicator_text,
                item['hash'],
                file_names,
                computer_names,
                local_addresses
            ]

            table.add_row(format_row(row))

        # After adding all rows, convert the table to a string
        table_str = table.get_string(fields=table.field_names)

        # Print the table
        print(table_str)

    else:
        print(f"\n{Fore.GREEN}{Style.BRIGHT}No malicious or suspicious indicators found.{Style.RESET_ALL}")

    print("\nProcessing complete.")
    print(f"\n\033[1;34mDashboard data exported to: \033[0;37m'\033[0;37m{full_path}\033[0;37m'\033[1;34m\033[0m")

    def format_detection_table(detections):
        """Format the detection information into a pretty table with complete hash values and colored formatting"""
        table = PrettyTable()
        headers = ["Type", "Detected Value", "Associated Filename", "Count"]
        table.field_names = [f"{Fore.CYAN}{Style.BRIGHT}{header}{Style.RESET_ALL}" for header in headers]
        table.align = "l"  # Left align text

        # Set fixed column widths
        table._max_width = {
            "Type": 15,
            "Detected Value": 65,
            "Associated Filename": 30,
            "Count": 10
        }

        # Enable word wrapping and borders
        table.hrules = True
        table.vrules = True

        def format_filenames(filenames):
            """Format filenames with proper wrapping and consistent yellow coloring"""
            formatted_parts = []
            for filename in filenames:
                # Handle each filename as a complete string first
                parts = []
                current_line = ""
                words = filename.split()

                for word in words:
                    # Check if adding this word would exceed the width
                    if len(current_line) + len(word) + 1 <= table._max_width["Associated Filename"]:
                        current_line = (current_line + " " + word).strip()
                    else:
                        if current_line:
                            parts.append(current_line)
                        current_line = word

                        # If single word is longer than max width, split it
                        while len(current_line) > table._max_width["Associated Filename"]:
                            parts.append(current_line[:table._max_width["Associated Filename"]])
                            current_line = current_line[table._max_width["Associated Filename"]:]

                if current_line:
                    parts.append(current_line)

                # Color each part individually and add to formatted_parts
                colored_parts = [f"{Fore.YELLOW}{Style.BRIGHT}{part}{Style.RESET_ALL}" for part in parts]
                formatted_parts.extend(colored_parts)

            return '\n'.join(formatted_parts)

        # Add rows for hash detections
        for hash_val, details in detections['hashes'].items():
            type_col = f"{Fore.LIGHTBLUE_EX}{Style.BRIGHT}● Hash{Style.RESET_ALL}"
            detected_val = hash_val  # Keep hash value without color
            filenames_col = format_filenames(details['filenames'])
            count_col = f"{Fore.GREEN}{Style.BRIGHT}{details['count']}{Style.RESET_ALL}"

            table.add_row([
                type_col,
                detected_val,
                filenames_col,
                count_col
            ])

        # Add rows for filename detections
        for proc_name, details in detections['filenames'].items():
            type_col = f"{Fore.LIGHTBLUE_EX}{Style.BRIGHT}● Process Name{Style.RESET_ALL}"
            detected_val = proc_name  # Keep process name without color
            filenames_col = format_filenames(details['filenames'])
            count_col = f"{Fore.GREEN}{Style.BRIGHT}{details['count']}{Style.RESET_ALL}"

            table.add_row([
                type_col,
                detected_val,
                filenames_col,
                count_col
            ])

        return table

    def remove_false_positives():
        try:
            # Get the dashboard filename
            filename = "SpectraShield_Dashboard.xlsx"
            full_path = os.path.abspath(filename)
            if not os.path.exists(filename):
                print(f"\n{Fore.RED}Error: Dashboard file '{filename}' not found.{Style.RESET_ALL}")
                return

            # Initialize detection tracking
            detections = {
                'hashes': {},    # {hash: {'filenames': set(), 'count': 0}}
                'filenames': {}  # {process_name: {'filenames': set(), 'count': 0}}
            }

            # Initialize sets to track unique values checked
            total_sha256 = 0
            total_md5 = 0
            total_filenames = 0

            # Check if data directory exists
            data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
            if not os.path.exists(data_dir):
                print(f"\n{Fore.RED}Error: 'data' directory not found.{Style.RESET_ALL}")
                return

            # Define paths for hash and process name files
            hash_file = os.path.join(data_dir, "list_of_false_positive_hashes.txt")
            process_file = os.path.join(data_dir, "list_of_false_positive_process_name.txt")

            # Check if required files exist
            if not os.path.exists(hash_file):
                print(f"\n{Fore.RED}Error: Hash list file not found at '{hash_file}'{Style.RESET_ALL}")
                return
            if not os.path.exists(process_file):
                print(f"\n{Fore.RED}Error: Process list file not found at '{process_file}'{Style.RESET_ALL}")
                return

            # Read hash values and process names from files
            with open(hash_file, 'r') as f:
                hashes = set(line.strip() for line in f if line.strip() and not line.startswith('#'))
            with open(process_file, 'r') as f:
                process_names = set(line.strip() for line in f if line.strip() and not line.startswith('#'))

            # Load the workbook
            wb = openpyxl.load_workbook(filename)
            ws = wb.active

            # Find the relevant column indices
            header_row = 1
            headers = {cell.value: cell.column for cell in ws[header_row]}

            sha256_col = headers.get("SHA256HashData")
            md5_col = headers.get("MD5HashData")
            filename_col = headers.get("FileName")
            sign_info_col = headers.get("sign_info")
            vendors_reported_col = headers.get("Vendors Reported")
            source_col = headers.get("Source")
            assigned_to_col = headers.get("Assigned_to")
            indicator_col = headers.get("Indicator")
            timestamp_col = headers.get("Report_timestamp")

            if not all([sha256_col, md5_col, filename_col, timestamp_col]):
                print(f"\n{Fore.RED}Error: Required columns not found in the dashboard.{Style.RESET_ALL}")
                return

            # Define font styles
            base_font = openpyxl.styles.Font(name='Calibri', size=11, color="D3D3D3")
            header_font = openpyxl.styles.Font(name='Calibri', size=11, bold=True)
            light_green_font = openpyxl.styles.Font(name='Calibri', size=11, color="92D050")
            red_bold_font = openpyxl.styles.Font(name='Calibri', size=11, color="FF0000", bold=True)
            yellow_font = openpyxl.styles.Font(name='Calibri', size=11, color="FFFF00")
            source_font = openpyxl.styles.Font(name='Calibri', size=11, color="87CEEB", underline='single')  # Light blue color
            assigned_to_font = openpyxl.styles.Font(name='Calibri', size=11, color="FFFFFF", bold=True)  # White bold

            # Store original Indicator column styles
            indicator_styles = {}
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=indicator_col)
                indicator_styles[row] = {
                    'font': copy.copy(cell.font) if cell.font else base_font,
                    'fill': copy.copy(cell.fill) if cell.fill else None,
                    'border': copy.copy(cell.border) if cell.border else None,
                    'alignment': copy.copy(cell.alignment) if cell.alignment else None,
                    'number_format': cell.number_format if cell.number_format else None
                }

            # Find the latest timestamp
            latest_timestamp = None
            for row in range(2, ws.max_row + 1):
                timestamp = ws.cell(row=row, column=timestamp_col).value
                if timestamp:
                    if latest_timestamp is None or timestamp > latest_timestamp:
                        latest_timestamp = timestamp

        # First pass: Count all entries before processing false positives
            for row in range(2, ws.max_row + 1):
                timestamp = ws.cell(row=row, column=timestamp_col).value
                if timestamp == latest_timestamp:  # Only count entries from the latest batch
                    sha256_value = str(ws.cell(row=row, column=sha256_col).value or '').strip()
                    md5_value = str(ws.cell(row=row, column=md5_col).value or '').strip()
                    filename_value = str(ws.cell(row=row, column=filename_col).value or '').strip()

                    if sha256_value:
                        total_sha256 += 1
                    if md5_value:
                        total_md5 += 1
                    if filename_value:
                        total_filenames += 1

            # Store row data and process false positives
            row_data = []
            for row in range(2, ws.max_row + 1):
                timestamp = ws.cell(row=row, column=timestamp_col).value
                sha256_value = str(ws.cell(row=row, column=sha256_col).value or '').strip()
                md5_value = str(ws.cell(row=row, column=md5_col).value or '').strip()
                filename_value = str(ws.cell(row=row, column=filename_col).value or '').strip()

                should_clear = False
                match_found = False

                # Only process rows with the latest timestamp
                if timestamp == latest_timestamp:
                    # Check and track hash matches
                    if sha256_value in hashes:
                        match_found = True
                        should_clear = True
                        if sha256_value not in detections['hashes']:
                            detections['hashes'][sha256_value] = {'filenames': set(), 'count': 0}
                        detections['hashes'][sha256_value]['filenames'].add(filename_value)
                        # Only increment count if we're actually removing this row
                        if should_clear:
                            detections['hashes'][sha256_value]['count'] += 1

                    if md5_value in hashes and not match_found:  # Only check MD5 if SHA256 didn't match
                        match_found = True
                        should_clear = True
                        if md5_value not in detections['hashes']:
                            detections['hashes'][md5_value] = {'filenames': set(), 'count': 0}
                        detections['hashes'][md5_value]['filenames'].add(filename_value)
                        # Only increment count if we're actually removing this row
                        if should_clear:
                            detections['hashes'][md5_value]['count'] += 1

                    # Check and track process name matches only if no hash matches were found
                    if not match_found:
                        for process_name in process_names:
                            if process_name.lower() in filename_value.lower():
                                should_clear = True
                                if process_name not in detections['filenames']:
                                    detections['filenames'][process_name] = {'filenames': set(), 'count': 0}
                                detections['filenames'][process_name]['filenames'].add(filename_value)
                                # Only increment count if we're actually removing this row and it's from the latest timestamp
                                if should_clear:
                                    detections['filenames'][process_name]['count'] += 1
                                break

                # If this row doesn't have the latest timestamp OR shouldn't be cleared, keep it
                if timestamp != latest_timestamp or not should_clear:
                    row_info = {
                        'values': [],
                        'hyperlinks': [],
                        'indicator_style': indicator_styles.get(row),
                        'height': ws.row_dimensions[row].height if row in ws.row_dimensions else None,
                        'is_empty': True
                    }

                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        value = cell.value
                        hyperlink = cell.hyperlink if hasattr(cell, 'hyperlink') else None
                        row_info['values'].append(value)
                        row_info['hyperlinks'].append(hyperlink)
                        if value is not None and str(value).strip() != '':
                            row_info['is_empty'] = False

                    row_data.append(row_info)

            # Sort the data - non-empty rows first
            row_data.sort(key=lambda x: x['is_empty'])

            # Write back the sorted data while preserving styles and hyperlinks
            for idx, row_info in enumerate(row_data, start=2):
                # Set row height if it was previously set
                if row_info['height'] is not None:
                    ws.row_dimensions[idx].height = row_info['height']

                # Write cell values and hyperlinks
                for col_idx, (value, hyperlink) in enumerate(zip(row_info['values'], row_info['hyperlinks']), start=1):
                    cell = ws.cell(row=idx, column=col_idx)
                    cell.value = value

                    # Apply appropriate styling based on column
                    if col_idx == indicator_col and row_info['indicator_style']:
                        # Preserve original Indicator column styling
                        cell.font = row_info['indicator_style']['font']
                        cell.fill = row_info['indicator_style']['fill']
                        cell.border = row_info['indicator_style']['border']
                        cell.alignment = row_info['indicator_style']['alignment']
                        cell.number_format = row_info['indicator_style']['number_format']
                    elif col_idx == filename_col:
                        cell.font = yellow_font
                    elif col_idx == vendors_reported_col:
                        if value and str(value).strip() == 'N/A':
                            cell.font = base_font
                        elif value and str(value).strip():
                            cell.font = red_bold_font
                        else:
                            cell.font = base_font
                    elif col_idx == sign_info_col and value and str(value).strip():
                        cell.font = light_green_font if str(value).strip() != 'N/A' else base_font
                    elif col_idx == source_col:  # Modified to apply to all cells in Source column
                        cell.font = source_font
                    elif col_idx == assigned_to_col:  # Modified to apply to all cells in Assigned_to column
                        cell.font = assigned_to_font
                    else:
                        cell.font = base_font

                    # Handle hyperlinks
                    if hasattr(cell, 'hyperlink'):
                        cell.hyperlink = None
                    if hyperlink is not None:
                        cell.hyperlink = hyperlink

            # Clear remaining rows
            for row in range(len(row_data) + 2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = None
                    if hasattr(cell, 'hyperlink'):
                        cell.hyperlink = None

            # Apply bold font to header row
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font

            # Save the workbook
            wb.save(filename)

            # Calculate total detections
            total_detections = sum(details['count'] for details in detections['hashes'].values()) + \
                            sum(details['count'] for details in detections['filenames'].values())

            # Calculate total entries checked (all entries, not just unique)
            total_entries = total_sha256 + total_md5 + total_filenames

            # Update the summary print section
            print(f"\n{Fore.CYAN}╔══ SpectraShield False Positive Detection Dashboard Update Summary ══╗{Style.RESET_ALL}")
            print(f"{Fore.CYAN}║{Style.RESET_ALL}")
            print(f"{Fore.CYAN}║{Style.RESET_ALL} {Fore.LIGHTBLUE_EX}{Style.BRIGHT}Entries checked before false positive removal:{Style.RESET_ALL}")
            print(f"{Fore.CYAN}║{Style.RESET_ALL} {Fore.YELLOW}{Style.BRIGHT}• Total SHA256 hashes checked:{Style.RESET_ALL} {total_sha256}")
            print(f"{Fore.CYAN}║{Style.RESET_ALL} {Fore.YELLOW}{Style.BRIGHT}• Total MD5 hashes checked:{Style.RESET_ALL} {total_md5}")
            print(f"{Fore.CYAN}║{Style.RESET_ALL} {Fore.YELLOW}{Style.BRIGHT}• Total filenames checked:{Style.RESET_ALL} {total_filenames}")
            print(f"{Fore.CYAN}║{Style.RESET_ALL}")
            print(f"{Fore.CYAN}║{Style.RESET_ALL} {Fore.YELLOW}{Style.BRIGHT}• Total entries checked:{Style.RESET_ALL} {total_entries}")
            print(f"{Fore.CYAN}║{Style.RESET_ALL}")
            if total_detections > 0:
                print(f"{Fore.CYAN}║{Style.RESET_ALL}{Fore.RED}{Style.BRIGHT}False Positive Detections:{Style.RESET_ALL}")
                print(f"{Fore.CYAN}║{Style.RESET_ALL}")
                table = format_detection_table(detections)
                for line in str(table).split('\n'):
                    print(f"{Fore.CYAN}║{Style.RESET_ALL} {line}")
            else:
                print(f"{Fore.CYAN}║{Style.RESET_ALL} No false positives detected.")

            print(f"{Fore.CYAN}║{Style.RESET_ALL}")
            print(f"{Fore.CYAN}║{Style.RESET_ALL} {Fore.YELLOW}{Style.BRIGHT}Total false positives removed:{Style.RESET_ALL} {total_detections}")
            print(f"{Fore.CYAN}║{Style.RESET_ALL}")
            print(f"{Fore.CYAN}║{Style.RESET_ALL} {Fore.LIGHTBLUE_EX}{Style.BRIGHT}Dashboard has been updated and saved to:{Style.RESET_ALL} {Fore.WHITE}'{full_path}'{Style.RESET_ALL}")
            print(f"{Fore.CYAN}╚{'═' * 69}╝{Style.RESET_ALL}")

        except Exception as e:
            print(f"\n{Fore.RED}Error occurred while removing false positives: {str(e)}{Style.RESET_ALL}")

    remove_false_positives()

    print("\nVirusTotal investigate scan completed.\n")

def Abuse_ch_API_Investigate(skip_prompt=False):
    if not skip_prompt:
        while True:
            confirm = input(f"\033[1;33mAre you sure you want to proceed with Abuse.ch API investigation? (yes/no): \033[0m").lower()
            if confirm == 'yes':
                break
            elif confirm == 'no':
                print("Returning to main menu.")
                return
            else:
                print("Invalid input. Please enter 'yes' or 'no'.")

    print("Proceeding with Abuse.ch API investigation...")

    # Configure alive_progress
    config_handler.set_global(spinner="waves", bar="smooth", unknown="waves", length=72, enrich_print=False)

    def get_terminal_width():
        return shutil.get_terminal_size().columns

    def configure_progress_bar():
        config_handler.set_global(spinner="waves", bar="smooth", unknown="waves", length=72, enrich_print=False, force_tty=True, stats=False)

    def format_progress_text(malicious_hashes, eta):
        terminal_width = get_terminal_width()
        available_width = terminal_width - 80
        malicious_text = f"{Fore.RED}{Style.BRIGHT}{malicious_hashes}{Style.RESET_ALL}"
        eta_text = f"{int(eta)}s"
        progress_text = f"Malicious: {malicious_text} | ETA: {eta_text}"
        padded_text = progress_text.ljust(available_width)
        return padded_text[:available_width]

    def query_abuse_ch_api(hash_value):
        command = f"curl -X POST -d 'query=get_info&hash={hash_value}' https://mb-api.abuse.ch/api/v1/"
        result = subprocess.run(command, shell=True, check=True, capture_output=True, text=True)
        return result.stdout

    def generate_filename():
        edt = pytz.timezone('US/Eastern')
        now = datetime.now(edt)
        return f"SpectraShield_Dashboard.xlsx"

    def get_current_timestamp():
        local_tz = get_localzone()
        current_time = datetime.now(local_tz)
        return current_time.strftime('%Y-%m-%d %H:%M:%S %Z')

    def format_results(json_result, hash_value, index, total_hashes):

        # Calculate the width needed for the counter
        counter_width = len(f"[{total_hashes}/{total_hashes}]")
        # Calculate the padding for all fields (counter width + 2 extra spaces)
        field_padding = counter_width + 2
        # Format the counter
        counter = f"[{index}/{total_hashes}]"
        # Create the base padding
        base_padding = ' ' * field_padding

        formatted_result = f"{counter:<{field_padding}}{Fore.YELLOW}{Style.BRIGHT}Results for hash:{Style.RESET_ALL} {Fore.BLUE}{hash_value}{Style.RESET_ALL}\n"
        if json_result.get("query_status") == "ok" and json_result.get("data"):
            data = json_result["data"][0]
            threat_classification = data.get("signature", "N/A")
            file_name = data.get("file_name", "N/A")
            sha256 = data.get("sha256_hash", "N/A")
            md5 = data.get("md5_hash", "N/A")

            formatted_result += f"{base_padding}{Fore.YELLOW}Popular Threat Classification:{Style.RESET_ALL} {Fore.WHITE}{threat_classification}{Style.RESET_ALL}\n"
            formatted_result += f"{base_padding}{Fore.YELLOW}File Name:{Style.RESET_ALL} {Fore.WHITE}{file_name}{Style.RESET_ALL}\n"
            formatted_result += f"{base_padding}{Fore.YELLOW}Last Analysis Stats:{Style.RESET_ALL} {{{Fore.RED}{Style.BRIGHT}malicious{Style.RESET_ALL}}}\n"
            formatted_result += f"{base_padding}{Fore.YELLOW}SHA256:{Style.RESET_ALL} {Fore.WHITE}{sha256}{Style.RESET_ALL}\n"
            formatted_result += f"{base_padding}{Fore.YELLOW}MD5:{Style.RESET_ALL} {Fore.WHITE}{md5}{Style.RESET_ALL}"
            return formatted_result, "Malicious", threat_classification, md5
        elif json_result.get("query_status") == "hash_not_found":
            formatted_result += f"{base_padding}{Fore.YELLOW}Status:{Style.RESET_ALL} {Fore.WHITE}No match found{Style.RESET_ALL}\n"
            formatted_result += f"{base_padding}{Fore.YELLOW}Details:{Style.RESET_ALL} {Fore.WHITE}The provided hash was not found in the Abuse.ch database.{Style.RESET_ALL}"
            return formatted_result, "Unknown", "N/A", "N/A"
        else:
            formatted_result += f"{base_padding}{Fore.YELLOW}Status:{Style.RESET_ALL} {Fore.WHITE}No results found{Style.RESET_ALL}"
            return formatted_result, "Unknown", "N/A", "N/A"

    def format_row(row):
        indicator, sha256_hash, file_names, computer_names, local_addresses = row

        indicator_color = Fore.LIGHTRED_EX
        indicator_text = f"{indicator_color}{Style.BRIGHT}●{Style.RESET_ALL} {indicator_color}{indicator}{Style.RESET_ALL}"

        return [
            indicator_text,
            sha256_hash,
            file_names,
            computer_names,
            local_addresses
        ]

    def main():

        total_hashes = 0
        malicious_hashes = 0

        while True:
            file_path = input(f"\033[1;33mPlease enter the full path to your CSV file: \033[0m").strip()
            if os.path.isfile(file_path):
                break
            print("The file does not exist. Please check the path and try again.")

        # Count total hashes
        print("Counting total hashes...")
        with open(file_path, 'r') as file:
            csv_reader = csv.DictReader(file)
            total_hashes = sum(1 for row in csv_reader)

        print(f"Total hashes to process: {total_hashes}")
        print("Processing hashes...")

        start_time = time.time()

        dashboard_data = []

        try:
            with open(file_path, 'r') as file:
                csv_reader = csv.DictReader(file)
                required_fields = ["SHA256HashData", "FileName", "ComputerName", "LocalAddressIP4", "aid", "cid"]
                if not all(field in csv_reader.fieldnames for field in required_fields):
                    print(f"Error: One or more required columns ({', '.join(required_fields)}) not found in the CSV file.")
                    return

                configure_progress_bar()

                with alive_bar(total_hashes, title="Processing hashes", enrich_print=False) as bar:
                    for index, row in enumerate(csv_reader, start=1):
                        hash_value = row["SHA256HashData"]
                        try:
                            result = query_abuse_ch_api(hash_value)
                            json_result = json.loads(result)
                            formatted_output, indicator, threat_classification, md5_hash = format_results(json_result, hash_value, index, total_hashes)

                            # Print above the progress bar
                            print(formatted_output)
                            print("")

                            # Add data to dashboard_data only if indicator is "Malicious"
                            if indicator == "Malicious":
                                malicious_hashes += 1
                                dashboard_data.append({
                                    "indicator": indicator,
                                    "hash": hash_value,
                                    "file_names": [row.get("FileName", "N/A")],
                                    "computer_names": [row.get("ComputerName", "N/A")],
                                    "local_addresses": [row.get("LocalAddressIP4", "N/A")],
                                    "threat_classification": threat_classification,
                                    "md5_hash": md5_hash,
                                    "aid": row.get("aid", "N/A"),
                                    "cid": row.get("cid", "N/A"),
                                    "file_path": row.get("FilePath", "N/A")
                                })

                        except json.JSONDecodeError:
                            print(f"Error: Unable to parse API response for hash {hash_value}")
                            print("")

                        # Calculate elapsed time and estimate the remaining time
                        elapsed_time = time.time() - start_time
                        estimated_total_time = (elapsed_time / index) * total_hashes if index > 0 else 0
                        eta = estimated_total_time - elapsed_time

                        # Update the progress bar with malicious hashes and ETA
                        progress_text = format_progress_text(malicious_hashes, eta)
                        bar.text(progress_text)
                        bar()

            end_time = time.time()
            elapsed_time = end_time - start_time

            # Create and populate the summary table
            table = PrettyTable()
            table.field_names = [
                f"{Fore.CYAN}{Style.BRIGHT}Indicator{Style.RESET_ALL}",
                f"{Fore.CYAN}{Style.BRIGHT}SHA256HashData{Style.RESET_ALL}",
                f"{Fore.CYAN}{Style.BRIGHT}FileName{Style.RESET_ALL}",
                f"{Fore.CYAN}{Style.BRIGHT}ComputerName{Style.RESET_ALL}",
                f"{Fore.CYAN}{Style.BRIGHT}LocalAddressIP4{Style.RESET_ALL}"
            ]
            # Set all column alignments to left
            for field in table.field_names:
                table.align[field] = "l"
            # Set header alignment to center
            table.header_align = "c"
            max_width = {
                "SHA256HashData": 70,
                "FileName": 30,
                "ComputerName": 30,
                "LocalAddressIP4": 20
            }
            for field, width in max_width.items():
                table.max_width[field] = width
            table.hrules = ALL  # Enable horizontal lines between rows

            for item in dashboard_data:
                file_names = "\n".join(item['file_names'])
                computer_names = "\n".join(item['computer_names'])
                local_addresses = "\n".join(item['local_addresses'])
                
                wrapped_file_names = "\n".join(["\n".join(textwrap.wrap(name, width=max_width["FileName"])) for name in file_names.split('\n')])
                wrapped_computer_names = "\n".join(["\n".join(textwrap.wrap(name, width=max_width["ComputerName"])) for name in computer_names.split('\n')])
                formatted_file_names = "\n".join([f"{Fore.YELLOW}{Style.BRIGHT}{name}{Style.RESET_ALL}" for name in wrapped_file_names.split('\n')])
                formatted_local_addresses = "\n".join([f"{Fore.GREEN}{Style.BRIGHT}{addr}{Style.RESET_ALL}" for addr in local_addresses.split('\n')])
                
                row = [
                    item['indicator'],
                    item['hash'],
                    formatted_file_names,
                    wrapped_computer_names,
                    formatted_local_addresses
                ]
                table.add_row(format_row(row))

            # Display the dashboard on terminal
            print("\n" + "=" * 45)
            print(f"{Fore.CYAN}{Style.BRIGHT}Result Summary:{Style.RESET_ALL}")
            print(f"* {Fore.YELLOW}{Style.BRIGHT}Status:{Style.RESET_ALL} Abuse.ch scan completed successfully")
            if malicious_hashes > 0:
                print(f"* {Fore.YELLOW}{Style.BRIGHT}Detection:{Style.RESET_ALL} {Fore.RED}{Style.BRIGHT}{malicious_hashes} Malicious Hash Detected!{Style.RESET_ALL}")
            else:

                print(f"* {Fore.YELLOW}{Style.BRIGHT}Detection:{Style.RESET_ALL} {Fore.GREEN}{Style.BRIGHT}No Malicious Hash Found!{Style.RESET_ALL}")
            print(f"* {Fore.YELLOW}{Style.BRIGHT}Scan option selected:{Style.RESET_ALL} Abuse.ch API")
            print(f"* {Fore.YELLOW}{Style.BRIGHT}Result:{Style.RESET_ALL} {Fore.RED}{Style.BRIGHT}{malicious_hashes}{Style.RESET_ALL}/{total_hashes}")
            print(f"* {Fore.YELLOW}{Style.BRIGHT}Elapsed time:{Style.RESET_ALL} {elapsed_time:.2f} seconds")
            print("=" * 45)

            print(table)

            # Check if the file already exists
            filename = generate_filename()

            # Get current timestamp
            current_time = get_current_timestamp()

            if os.path.exists(filename):
                # Load existing workbook
                wb = openpyxl.load_workbook(filename)
                ws = wb.active
            else:
                # Create new workbook
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Dashboard"

                # Write headers only if it's a new file
                headers = ["Status", "Assigned_to", "Verdict", "Report_timestamp", "sign_info", "Indicator", "Vendors Reported", "Severity", "Source", "ThreatClassification", "MD5HashData", "SHA256HashData", "FileName", "ComputerName", "LocalAddressIP4", "aid", "cid", "FilePath", "Additional Notes"]
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(name='Calibri', size=11, bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            # Add dropdown for Status column
            status_options = ["OPEN", "WORK_IN_PROGRESS", "SOC_VALIDATION", "WAITING_FOR_CUSTOMER", "CLOSED"]
            status_validation = DataValidation(type="list", formula1=f'"{",".join(status_options)}"', allow_blank=True)
            ws.add_data_validation(status_validation)

            # Find the first empty row after the header
            first_empty_row = 2
            while ws.cell(row=first_empty_row, column=1).value is not None:
                first_empty_row += 1

            # Define the light gray border
            light_gray_border = Border(
                left=Side(style='thin', color='444444'),
                right=Side(style='thin', color='444444'),
                top=Side(style='thin', color='444444'),
                bottom=Side(style='thin', color='444444')
            )

            # Define black fill
            black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

            # Add new data to the worksheet
            for index, item in enumerate(dashboard_data, start=first_empty_row):
                for col in range(1, 20):  # 20 columns
                    cell = ws.cell(row=index, column=col)
                    if col in [1, 2, 3, 4, 5, 6, 7, 8, 9]:  # Columns to center
                        cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                    elif col in [16, 17]:  # aid and cid columns
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    # Apply black background to all cells except the header row
                    if cell.row > 1:
                        cell.fill = black_fill
                    cell.border = light_gray_border

                    # Apply black background to all cells except the header row
                    cell.fill = black_fill

                ws.cell(row=index, column=1, value="OPEN")
                ws.cell(row=index, column=2, value="")  # Assigned_to (empty)
                ws.cell(row=index, column=3, value="")  # Verdict (empty)
                ws.cell(row=index, column=4, value=current_time)
                ws.cell(row=index, column=5, value="N/A") #sign_info
                ws.cell(row=index, column=6, value=item['indicator'])
                ws.cell(row=index, column=7, value="N/A")
                ws.cell(row=index, column=8, value="Critical")
                # Source with hyperlink
                source_cell = ws.cell(row=index, column=9, value="abuse.ch")
                source_cell.hyperlink = Hyperlink(target="https://bazaar.abuse.ch/browse/", ref="abuse.ch")
                source_cell.font = Font(name='Calibri', size=11, color="0000FF", underline="single")

                ws.cell(row=index, column=10, value=item['threat_classification'])
                ws.cell(row=index, column=11, value=item['md5_hash'])
                ws.cell(row=index, column=12, value=item['hash'])
                ws.cell(row=index, column=13, value="\n".join(item['file_names']))  # Join with newline
                ws.cell(row=index, column=14, value="\n".join(item['computer_names']))  # Join with newline
                ws.cell(row=index, column=15, value="\n".join(item['local_addresses']))  # Join with newline
                ws.cell(row=index, column=16, value=item['aid'])
                ws.cell(row=index, column=17, value=item['cid'])
                ws.cell(row=index, column=18, value=item['file_path'])

            # Adjusting column widths and row heights
            max_row = ws.max_row
            for col in range(1, ws.max_column + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                for row in range(1, max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                if col == 1:  # Status column
                    adjusted_width = min(max(25, max_length + 4), 35)
                elif col == 2:  # Assigned_to
                    adjusted_width = max(25, max_length + 4)
                elif col == 3:  # Verdict
                    adjusted_width = max(20, max_length + 4)
                elif col == 4:  # Report_Timestamp
                    adjusted_width = max(18, max_length + 4)
                elif col == 5:  # sign_info column
                    adjusted_width = min(max(20, max_length + 4), 25)
                elif col == 9:  # Source 
                    adjusted_width = min(max(15, max_length + 4), 20)
                elif col == 10:  # ThreatClassification
                    adjusted_width = min(max(40, max_length + 4), 50)
                elif col == 11:  # MD5HashData
                    adjusted_width = max(30, max_length + 4)
                elif col == 12:  # SHA256HashData
                    adjusted_width = max(65, max_length + 4)
                elif col == 13:  # FileName
                    adjusted_width = min(max(40, max_length + 4), 50)  # Cap at 50
                elif col == 14:  # ComputerName
                    adjusted_width = min(max(20, max_length + 4), 30)  # Cap at 30
                elif col == 15:  # LocalAddressIP4
                    adjusted_width = min(max(20, max_length + 4), 30)  # Cap at 30
                elif col == 16:  # aid column
                    adjusted_width = min(max(40, max_length + 4), 45)  # Cap at 40
                elif col == 17:  # cid column
                    adjusted_width = min(max(40, max_length + 4), 45)  # Cap at 40
                elif col == 18:  # FilePath column
                    adjusted_width = min(max(50, max_length + 4), 60)  # Cap at 60
                elif col == 19:  # Additional Notes
                    adjusted_width = 80
                else:
                    adjusted_width = min(max_length + 4, 45)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Add auto filter to all columns
            ws.auto_filter.ref = ws.dimensions

            # Apply validation to Status column (column A) for all rows with data
            last_row = ws.max_row
            status_validation.add(f'A2:A{last_row}')

            # Sort the data based on the "Report_timestamp" column
            data = list(ws.iter_rows(min_row=2, values_only=True))

            def safe_datetime(value):
                if value and isinstance(value, str):
                    try:
                        return datetime.strptime(value, '%Y-%m-%d %H:%M:%S %Z')
                    except ValueError:
                        return datetime.min
                return datetime.min

            sorted_data = sorted(data, key=lambda x: safe_datetime(x[3]), reverse=True) #timestamp is column 4

            # Clear the existing data (except the header)
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).value = None

            # Write the sorted data back to the worksheet
            for i, row_data in enumerate(sorted_data, start=2):
                for j, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=i, column=j, value=value)

                    # Reapply hyperlink to the "Source" column
                    if j == 9:  # Source column
                        if value == "abuse.ch":
                            cell.hyperlink = Hyperlink(target="https://bazaar.abuse.ch/browse/", ref="abuse.ch")
                            cell.font = Font(color="0000FF", underline="single")
                        elif value == "VirusTotal":
                            sha256_hash = row_data[11]  # SHA256 hash is in column 12 (index 11)
                            vt_url = f"https://www.virustotal.com/gui/file/{sha256_hash}"
                            cell.hyperlink = Hyperlink(target=vt_url, ref="VirusTotal")
                            cell.font = Font(color="0000FF", underline="single")

            # Reapply styles and validation to the sorted data
            for row in range(2, len(sorted_data) + 2):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if col in [1, 2, 3, 4, 5, 6, 7, 8, 9]:  # Columns to center
                        cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                    elif col in [16, 17]:  # aid and cid columns
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                    # Apply black background to all cells except the header row
                    if cell.row > 1:
                        cell.fill = black_fill
                    cell.border = light_gray_border

            # Reapply validation to Status column (column A) for all rows with data
            last_row = len(sorted_data) + 1
            status_validation.add(f'A2:A{last_row}')

            # Adjust row heights dynamically based on the wrapped content after sorting
            for row in range(1, last_row + 1):
                max_height = 0
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        text_lines = str(cell.value).count('\n') + 1
                        text_height = text_lines * 15  # Assuming 15 units per line of text
                        wrapped_text = textwrap.wrap(str(cell.value), width=int(ws.column_dimensions[get_column_letter(col)].width))
                        wrapped_height = len(wrapped_text) * 15
                        max_height = max(max_height, text_height, wrapped_height)
                ws.row_dimensions[row].height = max_height

            # After writing all the data and before saving the workbook, apply styles to the header row
            for col in range(1, 19):
                header_cell = ws.cell(row=1, column=col)
                header_cell.font = Font(bold=True)
                header_cell.alignment = Alignment(horizontal='center', vertical='center')
                header_cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                header_cell.border = light_gray_border

            # Add default font definition:;
            default_font = Font(name='Calibri', size=11, color='D3D3D3')

            def apply_default_font(ws):
                special_columns = {
                    "FileName": "FFFF00",  # Yellow
                    "Source": None,  # Has special underline formatting
                    "Assigned_to": "FFFFFF"  # White
                }

                for row in ws.iter_rows():
                    for cell in row:
                        if cell.row != 1:  # Skip the header row
                            header_value = ws.cell(row=1, column=cell.column).value
                            # Only apply default font if the column doesn't have special formatting
                            if header_value not in special_columns:
                                cell.font = default_font

            def format_vendors_reported_column(ws):
                # Find the "Vendors Reported" column
                vendors_reported_col = None
                header_row = 1

                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=header_row, column=col).value
                    if isinstance(cell_value, str) and cell_value.strip() == "Vendors Reported":
                        vendors_reported_col = col
                        break

                # If column is found, apply formatting
                if vendors_reported_col is not None:

                    na_style = Font(name="Calibri", size=11, color="D3D3D3")  # Light gray for N/A
                    value_style = Font(name="Calibri", size=11, color="FF0000", bold=True)  # Bold red for values

                    # Apply formatting to each cell in the column
                    for row in range(header_row + 1, ws.max_row + 1):
                        cell = ws.cell(row=row, column=vendors_reported_col)

                        # Handle different types of "N/A" values
                        cell_value = str(cell.value).strip() if cell.value is not None else ""

                        if cell_value.upper() in ["N/A", "NA", "N/A", "", "NONE"]:
                            cell.font = na_style
                        elif cell_value:  # If cell has any other value
                            cell.font = value_style

                    return True  # Indicates successful formatting
                return False  # Indicates column was not found

            def format_headers_with_assigned_to_bold(ws):
                for col in range(1, ws.max_column + 1):
                    header_cell = ws.cell(row=1, column=col)
                    # Default header formatting
                    header_cell.font = Font(name='Calibri', size=11, bold=True)
                    header_cell.alignment = Alignment(horizontal='center', vertical='center')
                    header_cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    header_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                            top=Side(style='thin'), bottom=Side(style='thin'))

                    # Extra bold for "Assigned_to" column
                    if header_cell.value == "Assigned_to":
                        header_cell.font = Font(name='Calibri', size=11, bold=True, color="000000")

            # Extra bold for "Assigned_to" column
            if header_cell.value == "Assigned_to":
                header_cell.font = Font(name='Calibri', size=11, bold=True, color="000000")  # Black color, bold

            def format_special_columns(ws):
                # Find the column indices
                filename_col = None
                indicator_col = None
                assigned_to_col = None
                source_col = None
                sign_info_col = None
                for col in range(1, ws.max_column + 1):
                    header_value = ws.cell(row=1, column=col).value
                    if header_value == "FileName":
                        filename_col = col
                    elif header_value == "Indicator":
                        indicator_col = col
                    elif header_value == "Assigned_to":
                        assigned_to_col = col
                    elif header_value == "Source":
                        source_col = col
                    elif header_value == "sign_info":
                        sign_info_col = col

                # Apply default light gray color to all data cells first
                for row in range(2, ws.max_row + 1):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.font = Font(name="Calibri", size=11, color="D3D3D3")

                # Then apply special formatting
                if filename_col is not None:
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=filename_col)
                        cell.font = Font(name="Calibri", size=11, color="FFFF00")  # Yellow

                if indicator_col is not None:
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=indicator_col)
                        if cell.value == "Malicious":
                            cell.font = Font(name="Calibri", size=11, color="FF0000", bold=True)  # Bold Red
                        elif cell.value == "Unknown":
                            cell.font = Font(name="Calibri", size=11, color="ADD8E6")  # Light Blue
                        elif cell.value == "Suspicious":
                            cell.font = Font(name="Calibri", size=11, color="FFFF00")  # Yellow

                if assigned_to_col is not None:
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=assigned_to_col)
                        cell.font = Font(name="Calibri", size=11, color="FFFFFF", bold=True)  # White and bold

                if source_col is not None:
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=source_col)
                        cell.font = Font(name="Calibri", size=11, color="D3D3D3", underline="single")

                if sign_info_col is not None:
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=sign_info_col)
                        if cell.value != "N/A":
                            cell.font = Font(name="Calibri", size=11, color="99FF99")  # Light Green 3
                        else:
                            cell.font = Font(name="Calibri", size=11, color="D3D3D3")

            apply_default_font(ws)
            format_special_columns(ws)
            format_headers_with_assigned_to_bold(ws)
            format_vendors_reported_column(ws)
            # Save the workbook
            wb.save(filename)

            print(f"\n\033[1;34mDashboard data exported to: \033[0;37m'\033[0;37m{os.path.abspath(filename)}\033[0;37m'\033[1;34m\033[0m")

        except FileNotFoundError:
            print(f"Error: CSV file '{file_path}' not found. Please check the file path and try again.")
        except subprocess.CalledProcessError:
            sad_face = f"""
            {Fore.RED}{Style.BRIGHT}
            (╯︵╰)
            {Style.RESET_ALL}
            """
            print(f"{sad_face}{Fore.RED}{Style.BRIGHT}Error{Style.RESET_ALL}: No internet connection detected. Please check your network connection and try again.\n")
        except Exception as e:
            print(f"An unexpected error occurred: {e}")

    if __name__ == "__main__":
        main()

    print("\nAbuse.ch investigate scan completed.\n")

def Custom_hash_list_investigate(skip_prompt=False):
    if not skip_prompt:
        while True:
            confirm = input(f"\033[1;33mAre you sure you want to proceed with Custom hash list search? (yes/no): \033[0m").lower()
            if confirm == 'yes':
                break
            elif confirm == 'no':
                print("Returning to main menu.")
                return
            else:
                print("Invalid input. Please enter 'yes' or 'no'.")
    print("\033[1;33;41mAPOLOGIES, THIS OPTION IS UNDER CONSTRUCTION...\033[0m")
    #print("Proceeding with Custom hash list search...")
    # Add your code for custom hash list search here
    #print("\nCustom hash list investigate scan complete.\n")

def Full_Scan_Investigate():
    print(f"{Fore.RED}{Style.BRIGHT}WARNING{Style.RESET_ALL}{Fore.WHITE}: Full Scan includes VirusTotal API usage, which has restrictions for business use.")
    print(Fore.CYAN + "Usage: This comprehensive scan utilizes all available investigation methods, including:")
    print("  - VirusTotal API (with usage restrictions)")
    print("  - Abuse.ch API")
    print("  - Custom hash list search")
    print("This option provides the most thorough analysis but may have legal implications for commercial use.")

    while True:
        confirm = input(f"\033[1;33mAre you sure you want to proceed with this option? (yes/no): \033[0m").lower()
        if confirm == 'yes':
            print("Proceeding with Full Scan Investigate...")
            VirusTotal_API_Investigate(skip_prompt=True)
            Abuse_ch_API_Investigate(skip_prompt=True)
            Custom_hash_list_investigate(skip_prompt=True)
            print("\nFull Scan complete.\n")
            break
        elif confirm == 'no':
            print("Returning to main menu.")
            return
        else:
            print("Invalid input. Please enter 'yes' or 'no'.")

def main():
    print_banner()

    while True:
        print("\n" + "=" * 43)
        print(f"{Fore.YELLOW}{Style.BRIGHT}           SPECTRA SHIELD MENU           {Style.RESET_ALL}")
        print("=" * 43)
        print("┌─────────────────────────────────────────┐")
        print("│  [1] VirusTotal API Investigation       │")
        print("│  [2] Abuse.ch API Investigation         │")
        print("│  [3] Custom Hash List Investigation     │")
        print("│  [4] Full Scan Investigation            │")
        print("│  [5] Exit                               │")
        print("└─────────────────────────────────────────┘")
        choice = input(f"\n\033[1;33mEnter your choice (1-5): \033[0m")

        if choice == '1':
            VirusTotal_API_Investigate()
        elif choice == '2':
            Abuse_ch_API_Investigate()
        elif choice == '3':
            Custom_hash_list_investigate()
        elif choice == '4':
            Full_Scan_Investigate()
        elif choice == '5':
            print("Exiting the program. Thank you for using Spectra_Shield! See you soon...")
            break
        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()
